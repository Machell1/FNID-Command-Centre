"""Registry register workbook tests."""

from pathlib import Path

from openpyxl import load_workbook


WORKBOOK_PATH = (
    Path(__file__).resolve().parents[1]
    / "fnid_portal"
    / "static"
    / "forms"
    / "registers"
    / "fnid_registry_registers_workbook.xlsx"
)


def test_registry_register_workbook_has_official_sheets():
    assert WORKBOOK_PATH.exists()

    workbook = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)

    assert workbook.sheetnames == [
        "Workbook Index",
        "Case File Submission",
        "Case Movement",
        "DCRR",
        "Investigator Index CR6",
        "Correspondence",
        "Major Crime Register",
    ]
    assert workbook["Case File Submission"]["A6"].value == "Case File Submission Register for Court"
    assert workbook["Case Movement"]["A6"].value == "Case File Movement Register"
    assert workbook["DCRR"]["A6"].value == "Divisional Case Report Register (DCRR)"
    assert workbook["Investigator Index CR6"]["A6"].value == "Investigator Index Card - CR 6"
    assert workbook["Correspondence"]["A6"].value == "Inward/Outward Correspondence Register"
    assert workbook["Major Crime Register"]["A6"].value == "Station Major Crime Register"


def test_registry_workflow_links_workbook(logged_in_client):
    registry_resp = logged_in_client.get("/unit/registry")
    assert registry_resp.status_code == 200
    assert b"/policy/registers/workbook/download" in registry_resp.data

    workbook_resp = logged_in_client.get("/policy/registers/workbook/download")
    assert workbook_resp.status_code == 200
    assert workbook_resp.data.startswith(b"PK")
